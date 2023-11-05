import os
from functools import cached_property
from utils import Log
from openpyxl import load_workbook


log = Log('CargillsDataset')

class CargillsDataset:
    @property
    def file_path(self) -> str:
        return os.environ.get('FILE_CARGILLS')

    @staticmethod
    def parse_row(ws, i_row):
        [row_num, description, unit_price_lkr, quantity, unit, discount_lkr, value_lkr] = [ws.cell(i_row, i_col).value for i_col in range(1, 8)]
        if not row_num:
            return None
        return dict(
            row_num=int(row_num),
            description=description,
            unit_price_lkr=float(unit_price_lkr),
            quantity=int(quantity),
            unit=unit,
            discount_lkr=float(discount_lkr),
            value_lkr=float(value_lkr)

        )

    @staticmethod
    def parse_worksheet(ws):
        i_row = 2
        data = []
        while True:
            datum = CargillsDataset.parse_row(ws, i_row)
            if not datum:
                break
            data.append(datum)
  
            i_row += 1
        return data
            

    @cached_property
    def data(self):
        data= {}
        wb = load_workbook(filename=self.file_path)
        for ws in wb.worksheets:
            date_id = ws.title

            if not date_id:
                continue
            data[date_id] = CargillsDataset.parse_worksheet(ws)
        return data

    @cached_property
    def description_to_date_to_price(self):
        idx = {}
        for date_id in self.data:
            for d in self.data[date_id]:
                description = d['description']
                unit_price_lkr = d['unit_price_lkr']
                if description not in idx:
                    idx[description] = {}
                idx[description][date_id] = unit_price_lkr
        return idx    